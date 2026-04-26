# -*- coding: utf-8 -*-
"""
Auto-SAXS — pseudo-Rg analysis (forced window; no qRg enforcement) with diagnostics

Additions in this version (v9):
- FIX: indentation error in DataSet.read_dotdat (previous v8 issue) corrected.
- Rg uncertainty (Rg_err) via slope error (scaled by chi2_red).
- Curvature test via weighted quadratic fit: ln I = c + m x + a x^2, x = q^2.
  * Reports a, its standard error se_a, and t = a/se_a.
  * a > 0 suggests a "smile"; a < 0 suggests a "frown".
- Safe CSV export (truncate to shortest q-grid across files).
"""

from __future__ import annotations

import csv
import math
import os
import re
from pathlib import Path
from typing import List, Tuple
import numpy as np

# =========================
# Key Parameters (edit here)
# =========================
DATA_DIR = r"C:\Users\E100104\OneDrive - RMIT University\DATA\2025\5Early aggregation\ML\6\analysis\selected"  # Data folder
# Font scaling for all plots
FONT_SCALE = 2.0
# Optional: directly specify the full analysis output folder (replaces DATA_DIR/analysis3).
# If empty, the script will use DATA_DIR/analysis3 as before.
OUTPUT_DIR = r"C:\Users\E100104\OneDrive - RMIT University\DATA\2025\5Early aggregation\ML\6\analysis\selected\analysis4"  # e.g. r"C:\path\to\my\outputs\analysis3"

# 统一 Guinier / ps-Rg / sigma 的 q² 区间
Q2_WIN_MIN = 0.001
Q2_WIN_MAX = 0.008
GUINIER_Q2_MIN = Q2_WIN_MIN   # Guinier overlay 用
GUINIER_Q2_MAX = Q2_WIN_MAX
PSFIT_Q2_MIN   = Q2_WIN_MIN   # 黄色拟合线可视化窗口
PSFIT_Q2_MAX   = Q2_WIN_MAX

OVERLAY_Y_EXPAND = 3.0        # guinier y 轴放大倍数

# 全局网格设计 —— 叠在一个大图里的布局
GRID_COLS = 5        # 每行几个 subplot
GRID_ROWS = None     # None = 自动行数；改成 11 / 15 即固定行数

DEFAULT_ROWS = 1     # 单页默认行数（仅作默认值，主程序会覆盖）
DEFAULT_COLS = 1
DEFAULT_TITLE_FORMAT = 'stem'  # 'stem' or 'full'

# log-SAXS grouped+offset 参数设置
GROUP_SIZE = 27       # 每 3 条曲线叠在一个 subplot；每条曲线做 vertical offset
# Layout B (log-saxs_g3_offset_T) 固定布局设置
LAYOUT_B_DPI = 300  # Layout B 图片分辨率（DPI）
LAYOUT_B_COLS = 1   # Layout B 的列数
LAYOUT_B_ROWS = 1   # Layout B 的行数
LAYOUT_B_SUBPLOT_HEIGHT = 5.0  # Layout B 每个 subplot 的高度（宽度保持默认 3.0）
LAYOUT_B_CURVE_COLOR = None  # Layout B 曲线颜色
LAYOUT_B_LINEWIDTH = 1       # Layout B 线条粗细
LAYOUT_B_X_MIN = 0.01          # Layout B x 轴最小值
LAYOUT_B_X_MAX = 0.4            # Layout B x 轴最大值
# =========================

# ---------- Utilities ----------
def sorted_nicely(items: List[str]) -> List[str]:
    def convert(text: str):
        return int(text) if text.isdigit() else text
    def alphanum_key(key: str):
        return [convert(c) for c in re.split(r"([0-9]+)", key)]
    return sorted(items, key=alphanum_key)

def _safe_isnan(x) -> bool:
    try:
        return math.isnan(x)
    except Exception:
        return False

def _clip_small(x: float, eps: float = 1e-20) -> float:
    return x if (x is not None and math.isfinite(x) and x > eps) else eps

# Invert a symmetric 3x3 matrix using adjugate / determinant
def _inv3(a11,a12,a13,a22,a23,a33):
    # matrix:
    # [a11 a12 a13]
    # [a12 a22 a23]
    # [a13 a23 a33]
    c11 =  a22*a33 - a23*a23
    c12 = -(a12*a33 - a23*a13)
    c13 =  a12*a23 - a22*a13
    c22 =  a11*a33 - a13*a13
    c23 = -(a11*a23 - a12*a13)
    c33 =  a11*a22 - a12*a12
    det = a11*c11 + a12*c12 + a13*c13
    if det == 0 or not math.isfinite(det):
        return None, None, None, None, None, None, None
    inv = (c11/det, c12/det, c13/det, c22/det, c23/det, c33/det, det)
    return inv

# ---------- Data container ----------
class DataSet:
    def __init__(self, dotdat: str):
        self.dotdat = dotdat
        self.q_array: List[float] = []
        self.qsq_array: List[float] = []
        self.i_array: List[float] = []
        self.i_raw_array: List[float] = []
        self.e_array: List[float] = []
        self.log_i_array: List[float] = []
        self.ln_i_array: List[float] = []
        self.i_qsq_array: List[float] = []
        self.guinier_model_array: List[float] = []
        self.tag: str = Path(dotdat).stem
        self.cycle_stats_array: List[list] = []

    def clean_negs(self) -> List[float]:
        out: List[float] = []
        for v in self.i_raw_array:
            if v is None or _safe_isnan(v) or v <= 0:
                out.append(float('nan'))
            else:
                out.append(float(v))
        self.i_array = out
        return self.i_array

    def log_intensity_calc(self) -> List[float]:
        out: List[float] = []
        for v in self.i_array:
            if v is None or _safe_isnan(v) or v <= 0:
                out.append(float('nan'))
            else:
                out.append(math.log10(v))
        self.log_i_array = out
        return self.log_i_array

    def ln_intensity_calc(self) -> List[float]:
        out: List[float] = []
        for v in self.i_array:
            if v is None or _safe_isnan(v) or v <= 0:
                out.append(float('nan'))
            else:
                out.append(math.log(v))
        self.ln_i_array = out
        return self.ln_i_array

    def qsq_calc(self) -> List[float]:
        self.qsq_array = [q * q for q in self.q_array]
        return self.qsq_array

    def i_qsq_calc(self) -> List[float]:
        out: List[float] = []
        for i, q2 in zip(self.i_array, self.qsq_array):
            if i is None or _safe_isnan(i) or i <= 0:
                out.append(float('nan'))
            else:
                out.append(i * q2)
        self.i_qsq_array = out
        return self.i_qsq_array

    def calculate_derivs(self):
        self.log_intensity_calc()
        self.ln_intensity_calc()
        self.qsq_calc()
        self.i_qsq_calc()

    def read_dotdat(self):
        self.q_array, self.i_raw_array, self.e_array = [], [], []
        with open(self.dotdat, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                # skip obvious headers or blank/short lines
                if not line.strip():
                    continue
                parts = re.split(r'[,\s]+', line.strip())
                
                if len(parts) < 2:
                    continue
                try:
                    q = float(parts[0])
                    i = float(parts[1])
                    e = float(parts[2]) if len(parts) > 2 else 1.0  # assign dummy error if missing
                  
                except ValueError:
                    # header tokens like "q  I  Err" will land here and be skipped
                    continue
                if not (math.isfinite(q) and math.isfinite(i) and math.isfinite(e)):
                    continue
                self.q_array.append(q)
                self.i_raw_array.append(i)
                self.e_array.append(e)
    
        # enforce strictly increasing q and drop any non-positive I for logs later
        zipped = sorted(zip(self.q_array, self.i_raw_array, self.e_array), key=lambda t: t[0])
        if not zipped:
            # no numeric data rows found; leave arrays empty
            self.q_array, self.i_raw_array, self.e_array = [], [], []
            self.i_array = []
            self.qsq_array = []
            return
        self.q_array, self.i_raw_array, self.e_array = map(list, zip(*zipped))
        self.clean_negs()
        self.qsq_calc()

# ---------- Analysis ----------
class AnalysisRun:
    def __init__(self):
        # Input / output
        self.root_path: str = ''
        self.analysis_path: str = ''
        
        self.bin_dq: float = float('nan')

        # File bookkeeping
        self.dat_list: List[str] = []
        self.dat_number: int = 0

        # Forced-window controls
        self.force_qsq_window: bool = True
        self.forced_qsq_lims: Tuple[float, float] = (Q2_WIN_MIN, Q2_WIN_MAX)

        # Record-keeping for forced-window provenance
        self.requested_qsq_lims: Tuple[float, float] = (Q2_WIN_MIN, Q2_WIN_MAX)
        self.q2_common_range: Tuple[float, float] = (float('nan'), float('nan'))
        self.forced_window_adjusted: bool = False
        self.forced_window_note: str = 'unchanged'
        self._warnings: List[str] = []

        # Ensemble storage
        self.ensemble_dat_names: List[str] = []
        self.ensemble_q_list: List[float] = []        # legacy
        self.ensemble_qsq_list: List[float] = []      # legacy
        self.ensemble_intensity_list: List[List[float]] = []
        self.ensemble_log_intensity_list: List[List[float]] = []
        self.ensemble_ln_intensity_list: List[List[float]] = []
        self.ensemble_i_qsq_list: List[List[float]] = []
        self.ensemble_guinier_model: List[List[float]] = []
        self.ensemble_stats: List[list] = []
        self._q_vectors: List[List[float]] = []       # for safe CSV export

    # ---- Paths ----
    def create_output_folder(self):
        # Allow overriding the output location via the module-level OUTPUT_DIR.
        # If OUTPUT_DIR is provided, use it directly as the analysis output folder.
        if OUTPUT_DIR and str(OUTPUT_DIR).strip():
            ap = Path(OUTPUT_DIR)
        else:
            ap = Path(self.root_path) / 'analysis3'
        ap.mkdir(parents=True, exist_ok=True)
        self.analysis_path = str(ap) + os.sep

        # Reset warnings log for this run
        try:
            with open(ap / 'WARNINGS.txt', 'w', encoding='utf-8') as f:
                f.write('Warnings (analysis3/WARNINGS.txt)\n')
        except Exception:
            pass


    def _emit_warning(self, msg: str):
        """Print and persist a warning message under analysis3/."""
        print(msg)
        self._warnings.append(msg)
        try:
            out = Path(self.analysis_path) / 'WARNINGS.txt'
            with open(out, 'a', encoding='utf-8') as f:
                f.write(msg + '\n')
        except Exception:
            pass

    # ---- Weighted least-squares with uncertainties ----
    @staticmethod
    def _wls_fit(xs: List[float], ys: List[float], sigmas: List[float]):
        n = len(xs)
        if n < 3:
            return (float('nan'),)*8

        w = [1.0 / (_clip_small(s)**2) for s in sigmas]

        Sw  = sum(w)
        Sx  = sum(wi * xi for wi, xi in zip(w, xs))
        Sy  = sum(wi * yi for wi, yi in zip(w, ys))
        Sxx = sum(wi * xi * xi for wi, xi in zip(w, xs))
        Sxy = sum(wi * xi * yi for wi, xi, yi in zip(w, xs, ys))

        Delta = Sw * Sxx - Sx * Sx
        if Delta == 0.0 or not math.isfinite(Delta):
            return (float('nan'),)*8

        m = (Sw * Sxy - Sx * Sy) / Delta
        c = (Sy * Sxx - Sx * Sxy) / Delta

        yhat = [m * x + c for x in xs]
        resid = [y - yh for y, yh in zip(ys, yhat)]

        chi2 = sum((r / _clip_small(s))**2 for r, s in zip(resid, sigmas))
        dof = max(1, n - 2)
        chi2_red = chi2 / dof

        ybar = Sy / Sw
        sst  = sum(wi * (yi - ybar) ** 2 for wi, yi in zip(w, ys))
        ssr  = sum(wi * (yi - yhi) ** 2 for wi, yi, yhi in zip(w, ys, yhat))
        R2_w = float('nan') if sst == 0.0 else 1.0 - (ssr / sst)

        # parameter variances (scaled by chi2_red)
        scale = chi2_red if math.isfinite(chi2_red) else 1.0
        var_m = scale * (Sw / Delta)
        var_c = scale * (Sxx / Delta)
        se_m = math.sqrt(var_m) if var_m >= 0 and math.isfinite(var_m) else float('nan')
        se_c = math.sqrt(var_c) if var_c >= 0 and math.isfinite(var_c) else float('nan')

        return m, c, chi2_red, R2_w, se_m, se_c, n, yhat

    @staticmethod
    def _wls_quad_fit(xs: List[float], ys: List[float], sigmas: List[float]):
        """Weighted quadratic regression y = c + m x + a x^2.
        Returns a, m, c, se_a, chi2_red, n
        """
        n = len(xs)
        if n < 4:
            return float('nan'), float('nan'), float('nan'), float('nan'), float('nan'), n

        w = [1.0 / (_clip_small(s)**2) for s in sigmas]

        Sw   = sum(w)
        Sx   = sum(wi * xi for wi, xi in zip(w, xs))
        Sx2  = sum(wi * xi*xi for wi, xi in zip(w, xs))
        Sx3  = sum(wi * (xi**3) for wi, xi in zip(w, xs))
        Sx4  = sum(wi * (xi**4) for wi, xi in zip(w, xs))

        Sy   = sum(wi * yi for wi, yi in zip(w, ys))
        Sxy  = sum(wi * xi * yi for wi, xi, yi in zip(w, xs, ys))
        Sx2y = sum(wi * (xi*xi) * yi for wi, xi, yi in zip(w, xs, ys))

        # Normal matrix (symmetric):
        # [Sw  Sx  Sx2]
        # [Sx  Sx2 Sx3]
        # [Sx2 Sx3 Sx4]
        inv = _inv3(Sw, Sx, Sx2, Sx2, Sx3, Sx4)
        if inv[0] is None:
            return float('nan'), float('nan'), float('nan'), float('nan'), float('nan'), n
        inv11, inv12, inv13, inv22, inv23, inv33, det = inv

        # params [c, m, a] = inv * [Sy, Sxy, Sx2y]
        c = inv11*Sy + inv12*Sxy + inv13*Sx2y
        m = inv12*Sy + inv22*Sxy + inv23*Sx2y
        a = inv13*Sy + inv23*Sxy + inv33*Sx2y

        # fitted values and chi2_red
        yhat = [c + m*x + a*x*x for x in xs]
        resid = [y - yh for y, yh in zip(ys, yhat)]
        chi2 = sum((r / _clip_small(s))**2 for r, s in zip(resid, sigmas))
        dof = max(1, n - 3)
        chi2_red = chi2 / dof

        # covariance = chi2_red * inv(X^T W X)^{-1} = chi2_red * inv (already computed)
        var_a = chi2_red * inv33
        se_a = math.sqrt(var_a) if (var_a is not None and math.isfinite(var_a) and var_a >= 0) else float('nan')

        return a, m, c, se_a, chi2_red, n

    @staticmethod
    def calc_rg(m: float) -> float:
        if m is None or not math.isfinite(m) or m >= 0.0:
            return float('nan')
        return math.sqrt(-3.0 * m)

    @staticmethod
    def calc_rg_err(Rg: float, se_m: float) -> float:
        if (Rg is None) or (se_m is None) or (not math.isfinite(Rg)) or (Rg <= 0) or (not math.isfinite(se_m)):
            return float('nan')
        return abs(3.0 / (2.0 * Rg)) * se_m

    # ---- Forced-window fit + diagnostics ----
    def lin_reg(self, cycle_data: DataSet):
        # sigma_lnI = Err / I
        sig_lnI = []
        for I, E in zip(cycle_data.i_array, cycle_data.e_array):
            if I and E and I > 0 and E > 0 and math.isfinite(I) and math.isfinite(E):
                sig_lnI.append(E / I)
            else:
                sig_lnI.append(float('nan'))

        lo, hi = self.forced_qsq_lims
        xs, ys, ss = [], [], []
        for q2, lnI, s in zip(cycle_data.qsq_array, cycle_data.ln_i_array, sig_lnI):
            if (q2 is None) or (lnI is None) or (s is None):
                continue
            if _safe_isnan(q2) or _safe_isnan(lnI) or _safe_isnan(s):
                continue
            if s <= 0:
                continue
            if lo <= float(q2) <= hi:
                xs.append(float(q2)); ys.append(float(lnI)); ss.append(float(s))

        if len(xs) < 3:
            cycle_data.cycle_stats_array.append([
                cycle_data.tag, float('nan'), float('nan'), float('nan'), float('nan'),
                float('nan'), float('nan'), float('nan'), float('nan'),
                float('nan'), float('nan'), float('nan'), 'insufficient points'
            ])


            cycle_data.guinier_model_array = [0.0] * len(cycle_data.qsq_array)
            return float('nan'), float('nan'), float('nan'), float('nan')

        # Linear fit
        m, c, chi2_red, R2_w, se_m, se_c, n, yhat = self._wls_fit(xs, ys, ss)
        Rg = self.calc_rg(m)
        Rg_err = self.calc_rg_err(Rg, se_m)

        # Quadratic curvature test
        a, m_q, c_q, se_a, chi2_red_q, n_q = self._wls_quad_fit(xs, ys, ss)
        t_a = abs(a / se_a) if (se_a is not None and math.isfinite(se_a) and se_a > 0 and math.isfinite(a)) else float('nan')

        # Verdict rules
        verdict = 'good'
        if not math.isfinite(Rg) or not math.isfinite(chi2_red):
            verdict = 'invalid'
        else:
            strong_curv = (t_a >= 3.0) if math.isfinite(t_a) else False
            mild_curv   = (2.0 <= t_a < 3.0) if math.isfinite(t_a) else False
            high_chi    = (chi2_red > 2.0) if math.isfinite(chi2_red) else False
            low_chi     = (chi2_red < 0.5) if math.isfinite(chi2_red) else False
            if strong_curv or high_chi:
                if math.isfinite(a) and a > 0:
                    verdict = 'smile curvature / wideness'
                elif math.isfinite(a) and a < 0:
                    verdict = 'frown curvature / background?'
                else:
                    verdict = 'poor fit'
            elif mild_curv or low_chi:
                verdict = 'borderline'
            else:
                verdict = 'good'

        # model only over used points
        used_set = set(xs)
        model = []
        for q2 in cycle_data.qsq_array:
            if (q2 in used_set) and math.isfinite(m) and math.isfinite(c):
                model.append(m * q2 + c)
            else:
                model.append(0.0)
        cycle_data.guinier_model_array = model

        # diagnostics
        if used_set and math.isfinite(Rg):
            max_q_used = math.sqrt(max(used_set))
            max_qRg = Rg * max_q_used
        else:
            max_qRg = float('nan')
        I0 = math.exp(c) if math.isfinite(c) else float('nan')

        # stats row: tag, m, c, Rg, Rg_err, chi2_red, max qRg, R2_w, I0, a, se_a, t_a, verdict
        # cycle_data.cycle_stats_array.append([
        #     cycle_data.tag, m, c, Rg, Rg_err, chi2_red, max_qRg, R2_w, I0, a, se_a, t_a, verdict
        # ])
        
        dq_bin = self.bin_dq
        sigma_q = dq_bin / math.sqrt(12.0) if math.isfinite(dq_bin) else float('nan')
        fwhm_q_quant = 2.0*math.sqrt(2.0*math.log(2.0))*sigma_q if math.isfinite(sigma_q) else float('nan')
        
        cycle_data.cycle_stats_array.append([
            cycle_data.tag, m, c, Rg, Rg_err, chi2_red, max_qRg, R2_w, I0, a, se_a, t_a, verdict,
            dq_bin, sigma_q, fwhm_q_quant])

        
        
        return m, c, chi2_red, Rg

    # ---- Ensemble and IO ----
    def collate_cycle(self, cycle_data: DataSet):
        self.ensemble_intensity_list.append(cycle_data.i_array)
        self.ensemble_log_intensity_list.append(cycle_data.log_i_array)
        self.ensemble_ln_intensity_list.append(cycle_data.ln_i_array)
        self.ensemble_i_qsq_list.append(cycle_data.i_qsq_array)
        self.ensemble_dat_names.append(cycle_data.tag)
        self.ensemble_q_list = cycle_data.q_array          # legacy
        self.ensemble_qsq_list = cycle_data.qsq_array      # legacy
        self.ensemble_guinier_model.append(cycle_data.guinier_model_array)
        self.ensemble_stats.append(cycle_data.cycle_stats_array)
        self._q_vectors.append(list(cycle_data.q_array))
    # numpy imported at module level; keep class namespace clean

    def _common_q_grid(self, all_q_lists: List[List[float]], n_points: int | None = None):
        # intersection of ranges
        q_min = max(min(qs) for qs in all_q_lists if qs)
        q_max = min(max(qs) for qs in all_q_lists if qs)
        if not (math.isfinite(q_min) and math.isfinite(q_max) and q_max > q_min):
            raise ValueError("No overlapping q-range across files.")
        # default: keep roughly the resolution of the shortest file
        if n_points is None:
            n_points = min(len(qs) for qs in all_q_lists if qs)
            n_points = max(50, n_points)  # don't go too tiny
        return np.linspace(q_min, q_max, n_points)
    
    @staticmethod
    def _interp_safe(x_src: List[float], y_src: List[float], x_new: np.ndarray) -> np.ndarray:
        # assumes x_src is strictly increasing, as ensured by read_dotdat()
        return np.interp(x_new, np.asarray(x_src, dtype=float), np.asarray(y_src, dtype=float))
    
    

    def crunch_ensemble(self, csv_type: str) -> List[list]:
        if not self.ensemble_intensity_list:
            return []

        # reference q: use the shortest q-grid across files
        min_len = min(len(qv) for qv in self._q_vectors)
        ref_q = next((qv for qv in self._q_vectors if len(qv) == min_len), self._q_vectors[0])[:min_len]
        q2_ref = [q * q for q in ref_q]

        def _slice_all(list_of_lists: List[List[float]]) -> List[List[float]]:
            return [lst[:min_len] for lst in list_of_lists]

        intens = _slice_all(self.ensemble_intensity_list)
        logI  = _slice_all(self.ensemble_log_intensity_list)
        lnI   = _slice_all(self.ensemble_ln_intensity_list)
        model = _slice_all(self.ensemble_guinier_model)

        if csv_type == 'saxs':
            rows = [['q'] + [f'{lbl} intensity' for lbl in self.ensemble_dat_names]]
            for i in range(min_len):
                row = [ref_q[i]]
                row.extend([vec[i] for vec in intens])
                rows.append(row)
            return rows

        if csv_type == 'log-saxs':
            rows = [['q'] + [f'{lbl} log(I)' for lbl in self.ensemble_dat_names]]
            for i in range(min_len):
                row = [ref_q[i]]
                row.extend([vec[i] for vec in logI])
                rows.append(row)
            return rows


        return []

    def write_csv(self, csv_type: str):
        rows = self.crunch_ensemble(csv_type)
        out = Path(self.analysis_path) / f'{csv_type}_{self.dat_number}.csv'
        with open(out, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f'Wrote {out}')

    def write_cycle_stats(self):
        out = Path(self.analysis_path) / 'Guinier_stats.txt'
        with open(out, 'w', encoding='utf-8') as f:
            # Run-level q^2-window diagnostics
            req_lo, req_hi = self.requested_qsq_lims
            use_lo, use_hi = self.forced_qsq_lims
            cmin, cmax = self.q2_common_range

            f.write(f"# Requested forced q^2 window: [{req_lo:.6g}, {req_hi:.6g}] A^-2\n")
            if math.isfinite(cmin) and math.isfinite(cmax):
                f.write(f"# Common-grid q^2 range:      [{cmin:.6g}, {cmax:.6g}] A^-2\n")
            f.write(f"# Used forced q^2 window:      [{use_lo:.6g}, {use_hi:.6g}] A^-2\n")
            if self.forced_window_adjusted:
                f.write(f"# WARNING: forced window adjusted: {self.forced_window_note}\n")
            elif self.forced_window_note and self.forced_window_note != 'unchanged':
                f.write(f"# NOTE: {self.forced_window_note}\n")

            f.write(
                'dat set, slope m, intercept c, Rg, Rg_err, chi2_red, max qRg, R2_w, I0, a, se_a, t_a, verdict, dq_bin, sigma_q, fwhm_q_quant\n'
            )
            for line in self.ensemble_stats:
                row = line[0]
                f.write(', '.join(str(x) for x in row) + '\n')
        print(f'Wrote {out}')

    def write_cycle_stats_xlsx(self):
        # Write Guinier_stats.xlsx with run-level forced-window diagnostics
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except Exception as e:
            self._emit_warning(f"WARNING: openpyxl not available; cannot write Guinier_stats.xlsx ({e}).")
            return

        out = Path(self.analysis_path) / 'Guinier_stats.xlsx'
        wb = Workbook()

        # Sheet 1: Stats
        ws = wb.active
        ws.title = 'Guinier_stats'

        headers = [
            'dat set', 'slope m', 'intercept c', 'Rg', 'Rg_err', 'chi2_red',
            'max qRg', 'R2_w', 'I0', 'a', 'se_a', 't_a', 'verdict',
            'dq_bin', 'sigma_q', 'fwhm_q_quant',
            'q2_req_lo', 'q2_req_hi', 'q2_used_lo', 'q2_used_hi', 'q2_common_lo', 'q2_common_hi',
            'window_adjusted', 'window_note'
        ]
        ws.append(headers)
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        req_lo, req_hi = self.requested_qsq_lims
        use_lo, use_hi = self.forced_qsq_lims
        cmin, cmax = self.q2_common_range
        note = self.forced_window_note
        adj = bool(self.forced_window_adjusted)

        for line in self.ensemble_stats:
            row = list(line[0])
            ws.append(row + [req_lo, req_hi, use_lo, use_hi, cmin, cmax, adj, note])

        ws.freeze_panes = 'A2'

        # Simple auto-width
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[col_letter].width = min(45, max(10, max_len + 2))

        # Sheet 2: RunInfo
        ws2 = wb.create_sheet('RunInfo')
        ws2.append(['key', 'value'])
        ws2['A1'].font = bold
        ws2['B1'].font = bold
        ws2.append(['DATA_DIR', str(self.root_path)])
        ws2.append(['n_files', int(self.dat_number)])
        ws2.append(['q2_req', f'[{req_lo:.6g}, {req_hi:.6g}]'])
        if math.isfinite(cmin) and math.isfinite(cmax):
            ws2.append(['q2_common', f'[{cmin:.6g}, {cmax:.6g}]'])
        else:
            ws2.append(['q2_common', 'nan'])
        ws2.append(['q2_used', f'[{use_lo:.6g}, {use_hi:.6g}]'])
        ws2.append(['window_adjusted', adj])
        ws2.append(['window_note', note])
        ws2.freeze_panes = 'A2'
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 70

        try:
            wb.save(out)
            print(f'Wrote {out}')
        except Exception as e:
            self._emit_warning(f"WARNING: failed to write Guinier_stats.xlsx ({e}).")


    def start(self):


        data_dir = Path(self.root_path)
        self.dat_list = sorted_nicely([str(p) for p in data_dir.glob('*.dat')])
        self.dat_number = len(self.dat_list)
        if self.dat_number == 0:
            print('No .dat files found.')
            return
        self.create_output_folder()
    
        datasets: List[DataSet] = []
        for dotdat in self.dat_list:
            ds = DataSet(dotdat)
            ds.read_dotdat()
            datasets.append(ds)
    
        # ▶ Build common q grid from overlap and interpolate all series
        all_q = [ds.q_array for ds in datasets if ds.q_array]
        try:
            q_common = self._common_q_grid(all_q)  # numpy array
        except ValueError as e:
            print(f"ERROR: {e}")
            return
        import numpy as _np
        self.bin_dq = float(_np.median(_np.diff(q_common)))
        q2_common = (q_common ** 2).tolist()
    
        for ds in datasets:
            # interpolate intensity and errors onto common grid
            I_interp  = self._interp_safe(ds.q_array, ds.i_array,  q_common)
            E_interp  = self._interp_safe(ds.q_array, ds.e_array,  q_common)
    
            # guard against negative/zero after interpolation for logs
            I_clipped = [val if (val is not None and math.isfinite(val) and val > 0) else float('nan')
                         for val in I_interp.tolist()]
            ds.q_array       = q_common.tolist()
            ds.qsq_array     = q2_common
            ds.i_raw_array   = I_interp.tolist()
            ds.i_array       = I_clipped
            ds.e_array       = E_interp.tolist()
            ds.calculate_derivs()  # rebuild log/ln, i·q² on the common grid

        # ▶ Optional: tighten Guinier window to common range if needed
        #    Important: the script may narrow/overwrite the requested window if it does not overlap
        #    the common q-grid (intersection across all files). Any such change is recorded and warned.
        if self.force_qsq_window:
            # snapshot the requested window BEFORE any tightening
            self.requested_qsq_lims = tuple(self.forced_qsq_lims)

            lo_req, hi_req = self.requested_qsq_lims
            q2_min, q2_max = q2_common[0], q2_common[-1]
            self.q2_common_range = (q2_min, q2_max)

            # only narrow; never expand past user's request
            new_lo = max(lo_req, q2_min)
            new_hi = min(hi_req, q2_max)

            # defaults
            self.forced_window_adjusted = False
            self.forced_window_note = 'unchanged'

            if new_hi - new_lo < 1e-6:
                # no meaningful overlap: fall back to full common grid
                self.force_qsq_window = True
                self.forced_qsq_lims = (q2_min, q2_max)
                self.forced_window_adjusted = True
                self.forced_window_note = (
                    f"requested [{lo_req:.6g}, {hi_req:.6g}] A^-2 has negligible overlap with common grid; "
                    f"using full common grid [{q2_min:.6g}, {q2_max:.6g}] A^-2"
                )
                self._emit_warning(
                    f"WARNING: Forced q^2 window requested [{lo_req:.6g}, {hi_req:.6g}] A^-2 has negligible overlap "
                    f"with common grid [{q2_min:.6g}, {q2_max:.6g}] A^-2; using full common grid instead."
                )
            else:
                self.forced_qsq_lims = (new_lo, new_hi)
                if (abs(new_lo - lo_req) > 0) or (abs(new_hi - hi_req) > 0):
                    self.forced_window_adjusted = True
                    self.forced_window_note = (
                        f"requested [{lo_req:.6g}, {hi_req:.6g}] A^-2 tightened to [{new_lo:.6g}, {new_hi:.6g}] A^-2 "
                        f"to fit common grid [{q2_min:.6g}, {q2_max:.6g}] A^-2"
                    )
                    self._emit_warning(
                        f"WARNING: Forced q^2 window requested [{lo_req:.6g}, {hi_req:.6g}] A^-2 was tightened to "
                        f"[{new_lo:.6g}, {new_hi:.6g}] A^-2 to match common grid [{q2_min:.6g}, {q2_max:.6g}] A^-2."
                    )

        # Now analyze and collate on the aligned grid
        for ds in datasets:
            self.lin_reg(ds)
            self.collate_cycle(ds)
    
        self.write_cycle_stats()
        self.write_cycle_stats_xlsx()
        for kind in ('saxs', 'log-saxs'):
            self.write_csv(kind)

# ---------- Main ----------

# =========================
# Paged plotting functions (from step1_plotting.py)
# =========================
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.image as mimage
import numpy as np
import csv
from pathlib import Path
import math
import re

# ---------- Plotting parameters ----------
# FONT_SCALE controls font size for all plots
FONT_SCALE = 2.0
base_font = float(matplotlib.rcParams.get('font.size', 10))
matplotlib.rcParams['font.size'] = base_font * FONT_SCALE
for key in ('axes.labelsize', 'axes.titlesize', 'xtick.labelsize', 'ytick.labelsize', 'legend.fontsize'):
    matplotlib.rcParams[key] = matplotlib.rcParams['font.size']




def _format_tag(tag: str, mode: str) -> str:
    if not isinstance(tag, str):
        return str(tag)
    if mode == 'full':
        return tag

    s = tag.strip()

    # 只在确实有 .dat 结尾时才去掉后缀
    if s.lower().endswith('.dat'):
        return s[:-4]

    # 只有当它看起来像路径（含 / 或 \）时，才用 Path 去掉路径部分
    if ('/' in s) or ('\\' in s):
        try:
            from pathlib import Path as _P
            return _P(s).stem
        except Exception:
            return s

    # 默认：原样返回（保留 EAN1.5_2 这种小数点）
    return s


def paged_plot_singlecurve(x, curves, plots_dir, rows=DEFAULT_ROWS, cols=DEFAULT_COLS, out_prefix='log-saxs', save_pdf=False, title_format=DEFAULT_TITLE_FORMAT, dpi=100, show_xlabel=False, linear_x=False):
    out_dir = Path(plots_dir) / 'saxs'
    out_dir.mkdir(parents=True, exist_ok=True)
    tags = list(curves.keys())
    n = len(tags)
    per_page = rows * cols
    pages = math.ceil(n / per_page)
    for p in range(pages):
        start = p * per_page
        end = min(n, start + per_page)
        page_tags = tags[start:end]
        for annot in [True, False]:
            suffix = '' if annot else '_noannot'
            fig, axes = plt.subplots(rows, cols, figsize=(3 * cols, 2.5 * rows), squeeze=False, sharey=False)
            for idx, tag in enumerate(page_tags):
                r = idx // cols
                c = idx % cols
                ax = axes[r][c]
                y = curves.get(tag)
                if y is None:
                    if annot:
                        ax.text(0.5, 0.5, f'MISSING: {tag}', ha='center', fontsize=matplotlib.rcParams['font.size'])
                    continue
                x_arr = np.asarray(x, dtype=float)
                y_arr = np.asarray(y, dtype=float)
                mask_q = (x_arr >= 0.01) & (x_arr <= 0.4)
                x_plot = x_arr[mask_q]
                y_plot = y_arr[mask_q]
                valid = np.isfinite(y_plot) & (y_plot != 0.0)
                x_plot = x_plot[valid]
                y_plot = y_plot[valid]
                if x_plot.size:
                    ax.plot(x_plot, y_plot, marker='.', linestyle='-', markersize=3, linewidth=2.0, color='blue')
                if annot:
                    ax.text(0.5, 0.02, _format_tag(tag, title_format), fontsize=matplotlib.rcParams.get('font.size'), ha='center', va='bottom', transform=ax.transAxes, bbox=dict(facecolor='white', alpha=0.1, edgecolor='none', boxstyle='round,pad=0.2'))
                if show_xlabel:
                    ax.set_xlabel('q', fontsize=matplotlib.rcParams.get('axes.labelsize'))
                if c == 0:
                    ax.set_ylabel('log I(q), a.u.', fontsize=matplotlib.rcParams.get('axes.labelsize'))
                ax.tick_params(axis='y', which='both', labelleft=False)
                try:
                    if y_plot.size:
                        arr = np.asarray(y_plot, dtype=float)
                        arr = arr[np.isfinite(arr)]
                        if arr.size:
                            y_min_curve = float(np.nanmin(arr))
                            y_max_curve = float(np.nanmax(arr))
                            yrange = y_max_curve - y_min_curve
                            if yrange <= 0.0:
                                pad = max(abs(y_max_curve) * 0.08, 0.4)
                                half = max(0.4, pad)
                                y0 = y_min_curve - half
                                y1 = y_max_curve + half
                            else:
                                pad = max(0.09 * yrange, 0.3)
                                min_span = 0.6
                                extra = (min_span - yrange) / 2.0 if yrange < min_span else 0.0
                                y0 = y_min_curve - (pad + extra)
                                y1 = y_max_curve + (pad + extra)
                            if np.isfinite(y0) and np.isfinite(y1) and (y0 < y1):
                                ax.set_ylim(y0, y1)
                except Exception:
                    pass
            total = rows * cols
            for j in range(len(page_tags), total):
                r = j // cols
                c = j % cols
                axes[r][c].axis('off')
            plt.subplots_adjust(wspace=0.0, hspace=0.0)
            for ax_row in axes:
                for ax in ax_row:
                    ax.set_xscale('linear' if linear_x else 'log')
                    ax.set_xlim(0.01, 0.4)
                    ax.tick_params(axis='x', which='both', labelbottom=show_xlabel)
            out = out_dir / f'{out_prefix}_page{p+1:02d}{suffix}.png'
            fig.savefig(out, dpi=dpi, bbox_inches='tight')
            plt.close(fig)



def quick_annotated_scatter(x, y, out_path, crystalline_present=None, fwhm_avg=None, title=None):
    plt.figure()
    plt.scatter(x, y, s=8)
    plt.xlabel("q")
    plt.ylabel("I(q)")
    if title:
        plt.title(title)
    # Add analysis keywords as annotation
    annotation = []
    if crystalline_present is not None:
        annotation.append(f"crystalline_present: {crystalline_present}")
    if fwhm_avg is not None:
        annotation.append(f"fwhm_avg: {fwhm_avg:.3f}")
    if annotation:
        #plt.gca().text(0.98, 0.02, '\n'.join(annotation), ha='right', va='bottom', transform=plt.gca().transAxes, fontsize=10, bbox=dict(facecolor='white', alpha=0.1, edgecolor='none'))
        plt.gca().text(0.98, 0.02, '\n'.join(annotation),
               ha='right', va='bottom', transform=plt.gca().transAxes, fontsize=10)

    plt.tight_layout()
    plt.savefig(out_path, dpi=120)
    plt.close()

def paged_plot_grouped_offset(
    x,
    curves,
    plots_dir,
    group_size: int = 3,
    rows: int = DEFAULT_ROWS,
    cols: int = DEFAULT_COLS,
    out_prefix: str = 'log-saxs_g3_offset',
    save_pdf: bool = False,
    title_format: str = DEFAULT_TITLE_FORMAT,
    dpi: int = 100,
    show_xlabel: bool = False,
    linear_x: bool = False,
    offset_mode: str = 'auto',          # 'auto' or 'fixed'
    offset_step: float | None = None,   # used when offset_mode == 'fixed'
    annot_style: str = 'textbox',       # 'textbox' | 'legend' | 'none'
    subplot_height: float | None = None,  # 每个 subplot 的高度（None 使用默认 2.5）
    curve_color: str | None = None,        # 曲线颜色（None 使用默认）
    linewidth: float | None = None,       # 线条粗细（None 使用默认 2.0）
    x_min: float | None = None,            # x 轴最小值（None 使用默认 0.01）
    x_max: float | None = None,            # x 轴最大值（None 使用默认 0.4）
    row_gap: float | None = None  # gap 高度（单位：英寸，推荐 0.3~1.5）
):
    """
    Plot log-SAXS curves with vertical offsets, grouped by `group_size` curves per subplot.
    - Each subplot contains up to `group_size` curves, each shifted by an offset.
    - Output is paged, analogous to paged_plot_singlecurve (but groups-as-subplots).
    """
    out_dir = Path(plots_dir) / 'saxs'
    out_dir.mkdir(parents=True, exist_ok=True)

    tags = list(curves.keys())
    if group_size < 1:
        group_size = 1

    # group tags: every `group_size` curves -> one subplot
    groups = [tags[i:i + group_size] for i in range(0, len(tags), group_size)]
    n_groups = len(groups)

    per_page = rows * cols
    pages = math.ceil(n_groups / per_page) if per_page > 0 else 1

    x_arr = np.asarray(x, dtype=float)
    # 使用自定义 x 轴范围或默认值
    x_min_val = x_min if x_min is not None else 0.01
    x_max_val = x_max if x_max is not None else 0.4
    mask_q = (x_arr >= x_min_val) & (x_arr <= x_max_val)
    x_base = x_arr[mask_q]

    def _auto_offset_step(y_arrays: list[np.ndarray]) -> float:
        # robust step based on the largest finite span among curves; ensure non-zero.
        spans = []
        for yy in y_arrays:
            if yy.size:
                v = yy[np.isfinite(yy)]
                if v.size:
                    spans.append(float(np.nanmax(v) - np.nanmin(v)))
        span = max(spans) if spans else 1.0
        if not math.isfinite(span) or span <= 0.0:
            span = 1.0
        # Slightly larger than span to visibly separate curves
        return 0.60 * span

    for p in range(pages):
        start = p * per_page
        end = min(n_groups, start + per_page)
        page_groups = groups[start:end]

        for annot in [True, False]:
            suffix = '' if annot else '_noannot'
            # if annot=False, force no text/legend
            annot_style_eff = annot_style if annot else 'none'


            # 每个 subplot 的高度（英寸）
            h = (subplot_height if (subplot_height is not None and math.isfinite(subplot_height) and subplot_height > 0) else 2.5)

            if row_gap is None or (not math.isfinite(row_gap)) or row_gap <= 0 or rows < 2:
                fig, axes = plt.subplots(rows, cols, figsize=(3 * cols, h * rows), squeeze=False, sharey=False)
            else:
                import matplotlib.gridspec as gridspec

                # 用 2*rows-1 行：subplot 行 + gap 行 + subplot 行 + ...
                n_gs_rows = 2 * rows - 1

                # gap 高度用“相对比例”表达：gap_ratio = row_gap / h
                gap_ratio = float(row_gap) / float(h)

                height_ratios = []
                for rr in range(rows):
                    height_ratios.append(1.0)            # subplot 行
                    if rr < rows - 1:
                        height_ratios.append(gap_ratio)  # gap 行

                fig = plt.figure(figsize=(3 * cols, h * rows + row_gap * (rows - 1)))
                gs = gridspec.GridSpec(n_gs_rows, cols, figure=fig, height_ratios=height_ratios)

                # axes[r][c] 映射到 GridSpec 的第 (2*r) 行（中间的 (2*r+1) 是 gap）
                axes = np.empty((rows, cols), dtype=object)
                for r in range(rows):
                    for c in range(cols):
                        axes[r][c] = fig.add_subplot(gs[2 * r, c])


            for idx, gtags in enumerate(page_groups):
                r = idx // cols
                c = idx % cols
                ax = axes[r][c]

                # gather curves for this group
                y_arrays = []
                x_arrays = []
                for tag in gtags:
                    y = curves.get(tag)
                    if y is None:
                        continue
                    y_arr = np.asarray(y, dtype=float)
                    y_sel = y_arr[mask_q]
                    valid = np.isfinite(y_sel) & (y_sel != 0.0)
                    x_plot = x_base[valid]
                    y_plot = y_sel[valid]
                    if x_plot.size and y_plot.size:
                        x_arrays.append(x_plot)
                        y_arrays.append(y_plot)

                if not y_arrays:
                    if annot:
                        ax.text(0.5, 0.5, f'MISSING: {gtags[0] if gtags else "group"}', ha='center', fontsize=matplotlib.rcParams['font.size'])
                    ax.axis('off')
                    continue

                # offset step
                if offset_mode == 'fixed' and (offset_step is not None) and math.isfinite(offset_step) and offset_step > 0:
                    step = float(offset_step)
                else:
                    step = _auto_offset_step(y_arrays)

                # plot with offsets
                ymins, ymaxs = [], []
                # 使用自定义颜色和线宽或默认值
                plot_color = curve_color if curve_color is not None else None
                plot_linewidth = linewidth if linewidth is not None else 2.0
                for j, (xx, yy) in enumerate(zip(x_arrays, y_arrays)):
                    off = j * step
                    yy_off = yy + off
                    ax.plot(xx, yy_off, marker='.', linestyle='-', markersize=3, 
                           linewidth=plot_linewidth, color=plot_color)
                    v = yy_off[np.isfinite(yy_off)]
                    if v.size:
                        ymins.append(float(np.nanmin(v)))
                        ymaxs.append(float(np.nanmax(v)))

                # annotation: show tag list
                if annot_style_eff == 'legend':
                    ax.legend([_format_tag(t, title_format) for t in gtags], loc='lower left', frameon=False)
                elif annot_style_eff == 'textbox':
                    lab = '\n'.join(_format_tag(t, title_format) for t in gtags)
                    ax.text(
                        0.02, 0.02, lab,
                        fontsize=matplotlib.rcParams.get('font.size'),
                        ha='left', va='bottom', transform=ax.transAxes,
                        
                    )

                if show_xlabel:
                    ax.set_xlabel('q', fontsize=matplotlib.rcParams.get('axes.labelsize'))
                if c == 0:
                    ax.set_ylabel('log I(q), a.u.', fontsize=matplotlib.rcParams.get('axes.labelsize'))

                ax.set_xscale('linear' if linear_x else 'log')
                ax.set_xlim(x_min_val, x_max_val)

                # y-limits based on plotted data
                if ymins and ymaxs:
                    y_min = min(ymins)
                    y_max = max(ymaxs)
                    yr = y_max - y_min
                    pad = max(0.1 * yr, 0.3)
                    top_pad = 1.8 * pad      # 上方留白加大（1.2~2.0 之间试）
                    bottom_pad = 1.0 * pad   # 下方留白保持不变（也可改 0.6*pad）
                    ax.set_ylim(y_min - bottom_pad, y_max + top_pad)

                # keep clean: match existing style (no y tick labels by default)
                ax.tick_params(axis='y', which='both', labelleft=False)
                ax.tick_params(axis='x', which='both', labelbottom=show_xlabel)

            # turn off unused subplots
            total = rows * cols
            for j in range(len(page_groups), total):
                rr = j // cols
                cc = j % cols
                axes[rr][cc].axis('off')

            # 使用 row_gap 控制两行 subplot 之间的垂直间距；
            # 如果 row_gap 为 None，则退回到原来的紧凑布局（hspace=0.0）
            hspace_val = row_gap if (row_gap is not None and row_gap >= 0.0) else 0.0
            plt.subplots_adjust(wspace=0.0, hspace=hspace_val)

            out = out_dir / f'{out_prefix}_page{p+1:02d}{suffix}.png'
            fig.savefig(out, dpi=dpi, bbox_inches='tight')
            plt.close(fig)


def read_table(csv_path):
    csv_path = Path(csv_path)
    with csv_path.open('r', newline='') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        raise ValueError(f'Empty CSV: {csv_path}')
    headers = rows[0]
    data = rows[1:]
    x = [float(r[0]) for r in data]
    cols = list(zip(*data)) if data else []
    result = {}
    for col_idx in range(1, len(headers)):
        header = headers[col_idx]
        clean = header.replace(' log(I) fit', '').replace(' log(I)', '')
        tag = clean.strip()
        vals = [float(v) if v != '' else float('nan') for v in cols[col_idx]]
        result[tag] = vals
    return {'x': x, 'data': result}

# ---------- Main ----------
if __name__ == '__main__':
    run = AnalysisRun()
    run.root_path = DATA_DIR
    run.force_qsq_window = True   # 是否启用强制窗口（区间本身在 __init__ 里统一定义）
    print(f"Reading .dat files from: {run.root_path}")
    # Determine where analysis outputs will be written (can be overridden by OUTPUT_DIR)
    if OUTPUT_DIR and str(OUTPUT_DIR).strip():
        analysis_dir = Path(OUTPUT_DIR)
    else:
        analysis_dir = Path(run.root_path) / 'analysis3'
    print(f"Outputs will be saved under: {analysis_dir}")
    run.start()

    # After analysis and CSV export, run paged plotting using generated CSVs
    plots_dir = analysis_dir / 'plots'
    logsaxs_csv = next(analysis_dir.glob('log-saxs_*.csv'), None)
    if logsaxs_csv:
        l = read_table(logsaxs_csv)
        x_l = l['x']
        log_curves = l['data']

        # 计算总的行数和列数：保证 9 列，行数足够容纳所有数据
        n_tags = len(log_curves)
        cols = GRID_COLS
        if GRID_ROWS is None:
            # 自动最小行数
            rows = (n_tags + cols - 1) // cols
        else:
            # 使用用户指定的行数，但保证能容纳所有 subplot
            rows = GRID_ROWS
            if rows * cols < n_tags:
                rows = (n_tags + cols - 1) // cols

        # log-SAXS 图：保持原来逻辑，但同样用 rows/cols
        paged_plot_singlecurve(
            x_l, log_curves, plots_dir,
            rows=rows, cols=cols,
            out_prefix='log-saxs',
            save_pdf=False,
            dpi=100,
            show_xlabel=False,
            linear_x=False
        )


        # ②b log-SAXS grouped+offset（每 GROUP_SIZE 条曲线叠在一个 subplot；每条曲线做 vertical offset）
        n_groups = (n_tags + GROUP_SIZE - 1) // GROUP_SIZE

        # 默认：把"原来一行 9 个图"变为"一行 3 个图"（即 cols // GROUP_SIZE），行数自动补足（或沿用 GRID_ROWS）
        group_cols = max(1, cols // GROUP_SIZE)
        if GRID_ROWS is None:
            group_rows = (n_groups + group_cols - 1) // group_cols
        else:
            group_rows = GRID_ROWS
            if group_rows * group_cols < n_groups:
                group_rows = (n_groups + group_cols - 1) // group_cols

        # Layout A：group_rows × group_cols
        paged_plot_grouped_offset(
            x_l, log_curves, plots_dir,
            group_size=GROUP_SIZE,
            rows=group_rows, cols=group_cols,
            out_prefix='log-saxs_g3_offset',
            save_pdf=False,
            dpi=100,
            show_xlabel=False,
            linear_x=False,
            annot_style='textbox',
            offset_mode='auto'             
        )

        # Layout B：固定布局（6 列 2 行）
        paged_plot_grouped_offset(
            x_l, log_curves, plots_dir,
            group_size=GROUP_SIZE,
            rows=LAYOUT_B_ROWS, cols=LAYOUT_B_COLS,
            out_prefix='log-saxs_g3_offset_T',
            save_pdf=False,
            dpi=LAYOUT_B_DPI,
            show_xlabel=False,
            linear_x=False,
            annot_style='textbox',
            offset_mode='auto',
            row_gap=0,
            subplot_height=LAYOUT_B_SUBPLOT_HEIGHT,
            curve_color=LAYOUT_B_CURVE_COLOR,
            linewidth=LAYOUT_B_LINEWIDTH,
            x_min=LAYOUT_B_X_MIN,
            x_max=LAYOUT_B_X_MAX
        )
        print('Plots saved to', plots_dir)

        # Example: quick figure generation for Buf_2.5mg_quick.png with analysis keywords
        buf_x = x_l
        # Debug: print available keys in log_curves
        print(f"Available keys in log_curves: {list(log_curves.keys())}")
        
        # Find a matching key (case-insensitive search for 'buf')
        buf_key = None
        for key in log_curves.keys():
            if 'buf' in key.lower():
                buf_key = key
                break
        
        if buf_key:
            buf_y = log_curves[buf_key]
            crystalline_present = 'yes'
            fwhm_avg = 0.123
            quick_annotated_scatter(
                buf_x, buf_y,
                plots_dir / f'{buf_key}_quick.png',
                crystalline_present=crystalline_present,
                fwhm_avg=fwhm_avg,
                title=buf_key
            )
        else:
            print("Warning: No 'buf' key found in log_curves. Skipping quick scatter plot.")

        # Combine annotated and non-annotated overlays for saxs folder
        from pathlib import Path
        import matplotlib.image as mimage
        import numpy as np
        # saxs folder
        saxs_dir = plots_dir / 'saxs'
        saxs_annot_files = sorted(saxs_dir.glob('log-saxs_page*_annot.png'))
        saxs_noannot_files = sorted(saxs_dir.glob('log-saxs_page*_noannot.png'))
        if saxs_annot_files:
            arrays = [mimage.imread(str(f)) for f in saxs_annot_files]
            stacked = np.vstack(arrays)
            mimage.imsave(str(saxs_dir / 'log-saxs_all_annot.png'), stacked)
        if saxs_noannot_files:
            arrays = [mimage.imread(str(f)) for f in saxs_noannot_files]
            stacked = np.vstack(arrays)
            mimage.imsave(str(saxs_dir / 'log-saxs_all_noannot.png'), stacked)

    else:
        print('Could not find all required CSVs for plotting in analysis3/.')

